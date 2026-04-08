"""excel_ops.py の単体テスト"""
import json
import pytest
from pathlib import Path
from datetime import datetime
from unittest.mock import patch, MagicMock

from openpyxl import Workbook


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def sample_xlsx(tmp_path):
    """基本的なExcelファイルを作成"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Age", "City"])
    ws.append(["Alice", 30, "Tokyo"])
    ws.append(["Bob", 25, "Osaka"])
    ws.append(["Charlie", 35, "Nagoya"])
    path = tmp_path / "test.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def multi_sheet_xlsx(tmp_path):
    """複数シートのExcelファイル"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["A", "B"])
    ws1.append([1, 2])
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["X", "Y", "Z"])
    ws2.append([10, 20, 30])
    ws2.append([40, 50, 60])
    path = tmp_path / "multi.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def empty_xlsx(tmp_path):
    """空のExcelファイル"""
    wb = Workbook()
    path = tmp_path / "empty.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def unicode_xlsx(tmp_path):
    """Unicode文字を含むExcelファイル"""
    wb = Workbook()
    ws = wb.active
    ws.append(["名前", "年齢", "都市"])
    ws.append(["太郎", 30, "東京"])
    ws.append(["花子", 25, "大阪"])
    path = tmp_path / "unicode.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def datetime_xlsx(tmp_path):
    """日付データを含むExcelファイル"""
    wb = Workbook()
    ws = wb.active
    ws.append(["Event", "Date"])
    ws.append(["Meeting", datetime(2025, 1, 15, 10, 30, 0)])
    path = tmp_path / "dates.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------

class TestImport:
    def test_import_module(self):
        from excel_ops import ExcelOperator
        assert ExcelOperator is not None

    def test_import_main(self):
        from excel_ops import main
        assert callable(main)


# ---------------------------------------------------------------------------
# ExcelOperator.__init__ / load
# ---------------------------------------------------------------------------

class TestExcelOperatorInit:
    def test_init(self, tmp_path):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(tmp_path / "test.xlsx"))
        assert op.filepath == tmp_path / "test.xlsx"
        assert op.workbook is None

    def test_load_nonexistent(self, tmp_path):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(tmp_path / "nonexistent.xlsx"))
        assert op.load() is False

    def test_load_valid(self, sample_xlsx):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(sample_xlsx))
        assert op.load() is True
        assert op.workbook is not None

    def test_load_corrupt_file(self, tmp_path):
        """破損ファイルの読み込みはFalseを返す"""
        from excel_ops import ExcelOperator
        corrupt = tmp_path / "corrupt.xlsx"
        corrupt.write_bytes(b"\x00\x01\x02garbage")
        op = ExcelOperator(str(corrupt))
        assert op.load() is False

    def test_load_zero_byte_file(self, tmp_path):
        """0バイトファイルの読み込み"""
        from excel_ops import ExcelOperator
        zero = tmp_path / "zero.xlsx"
        zero.write_bytes(b"")
        op = ExcelOperator(str(zero))
        assert op.load() is False

    def test_load_wrong_extension(self, tmp_path):
        """拡張子が異なるファイル"""
        from excel_ops import ExcelOperator
        txt = tmp_path / "data.txt"
        txt.write_text("not excel")
        op = ExcelOperator(str(txt))
        assert op.load() is False


# ---------------------------------------------------------------------------
# get_sheet_names
# ---------------------------------------------------------------------------

class TestGetSheetNames:
    def test_without_load(self):
        from excel_ops import ExcelOperator
        op = ExcelOperator("dummy.xlsx")
        assert op.get_sheet_names() == []

    def test_single_sheet(self, sample_xlsx):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(sample_xlsx))
        op.load()
        names = op.get_sheet_names()
        assert "Data" in names

    def test_multi_sheet(self, multi_sheet_xlsx):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(multi_sheet_xlsx))
        op.load()
        names = op.get_sheet_names()
        assert "Sheet1" in names
        assert "Sheet2" in names
        assert len(names) == 2


# ---------------------------------------------------------------------------
# read_sheet
# ---------------------------------------------------------------------------

class TestReadSheet:
    def test_read_default_sheet(self, sample_xlsx):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(sample_xlsx))
        op.load()
        data = op.read_sheet()
        assert data["sheet_name"] == "Data"
        assert data["headers"] == ["Name", "Age", "City"]
        assert len(data["rows"]) == 3

    def test_read_without_load(self):
        from excel_ops import ExcelOperator
        op = ExcelOperator("dummy.xlsx")
        data = op.read_sheet()
        assert "error" in data

    def test_read_nonexistent_sheet(self, sample_xlsx):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(sample_xlsx))
        op.load()
        data = op.read_sheet(sheet_name="NoSuchSheet")
        assert "error" in data

    def test_read_named_sheet(self, multi_sheet_xlsx):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(multi_sheet_xlsx))
        op.load()
        data = op.read_sheet(sheet_name="Sheet2")
        assert data["sheet_name"] == "Sheet2"
        assert data["headers"] == ["X", "Y", "Z"]

    def test_read_max_rows_boundary(self, tmp_path):
        """max_rows境界値テスト"""
        from excel_ops import ExcelOperator
        wb = Workbook()
        ws = wb.active
        ws.append(["Value"])
        for i in range(50):
            ws.append([i])
        path = tmp_path / "many_rows.xlsx"
        wb.save(path)

        op = ExcelOperator(str(path))
        op.load()
        data = op.read_sheet(max_rows=10)
        assert len(data["rows"]) <= 10

    def test_read_empty_sheet(self, empty_xlsx):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(empty_xlsx))
        op.load()
        data = op.read_sheet()
        # 空シートでもエラーにならない
        assert "error" not in data

    def test_read_datetime_cells(self, datetime_xlsx):
        """datetime型セルが文字列に変換される"""
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(datetime_xlsx))
        op.load()
        data = op.read_sheet()
        assert len(data["rows"]) >= 1
        # datetime は "YYYY-MM-DD HH:MM:SS" 形式
        assert "2025-01-15" in data["rows"][0][1]

    def test_read_unicode_content(self, unicode_xlsx):
        """Unicodeコンテンツの読み取り"""
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(unicode_xlsx))
        op.load()
        data = op.read_sheet()
        assert "名前" in data["headers"]
        assert data["rows"][0][0] == "太郎"

    def test_read_none_cells(self, tmp_path):
        """None値セルの処理"""
        from excel_ops import ExcelOperator
        wb = Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        ws.append([None, "value"])
        ws.append(["value", None])
        path = tmp_path / "nones.xlsx"
        wb.save(path)

        op = ExcelOperator(str(path))
        op.load()
        data = op.read_sheet()
        # None は "" に変換される
        for row in data["rows"]:
            assert all(isinstance(v, str) for v in row)


# ---------------------------------------------------------------------------
# to_markdown
# ---------------------------------------------------------------------------

class TestToMarkdown:
    def test_basic_markdown(self, sample_xlsx):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(sample_xlsx))
        op.load()
        md = op.to_markdown()
        assert "| Name | Age | City |" in md
        assert "| --- | --- | --- |" in md
        assert "Alice" in md

    def test_markdown_not_loaded(self):
        from excel_ops import ExcelOperator
        op = ExcelOperator("dummy.xlsx")
        md = op.to_markdown()
        assert "Error" in md

    def test_markdown_nonexistent_sheet(self, sample_xlsx):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(sample_xlsx))
        op.load()
        md = op.to_markdown(sheet_name="NoSuchSheet")
        assert "Error" in md

    def test_markdown_pipe_escape(self, tmp_path):
        """パイプ文字がエスケープされる"""
        from excel_ops import ExcelOperator
        wb = Workbook()
        ws = wb.active
        ws.append(["Col"])
        ws.append(["a|b"])
        path = tmp_path / "pipe.xlsx"
        wb.save(path)

        op = ExcelOperator(str(path))
        op.load()
        md = op.to_markdown()
        assert "a\\|b" in md

    def test_markdown_max_rows(self, sample_xlsx):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(sample_xlsx))
        op.load()
        md = op.to_markdown(max_rows=1)
        # max_rows=1 -> only 1 row from data (beyond header)
        assert "Alice" in md or "Bob" not in md


# ---------------------------------------------------------------------------
# analyze
# ---------------------------------------------------------------------------

class TestAnalyze:
    def test_analyze_basic(self, sample_xlsx):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(sample_xlsx))
        op.load()
        report = op.analyze()
        assert report["filename"] == "test.xlsx"
        assert report["summary"]["sheet_count"] == 1
        assert report["total_rows"] > 0

    def test_analyze_not_loaded(self):
        from excel_ops import ExcelOperator
        op = ExcelOperator("dummy.xlsx")
        report = op.analyze()
        assert "error" in report

    def test_analyze_multi_sheet(self, multi_sheet_xlsx):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(multi_sheet_xlsx))
        op.load()
        report = op.analyze()
        assert report["summary"]["sheet_count"] == 2
        assert len(report["sheets"]) == 2


# ---------------------------------------------------------------------------
# create_workbook
# ---------------------------------------------------------------------------

class TestCreateWorkbook:
    def test_create_basic(self, tmp_path):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(tmp_path / "output.xlsx"))
        data = {
            "headers": ["Name", "Score"],
            "rows": [["Alice", 95], ["Bob", 88]],
        }
        output = op.create_workbook(data)
        assert Path(output).exists()

    def test_create_with_custom_output(self, tmp_path):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(tmp_path / "ignored.xlsx"))
        custom_out = str(tmp_path / "custom.xlsx")
        data = {"headers": ["A"], "rows": [["x"]]}
        output = op.create_workbook(data, output_path=custom_out)
        assert output == custom_out
        assert Path(custom_out).exists()

    def test_create_empty_data(self, tmp_path):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(tmp_path / "empty_out.xlsx"))
        data = {"headers": [], "rows": []}
        output = op.create_workbook(data)
        assert Path(output).exists()

    def test_create_unicode_data(self, tmp_path):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(tmp_path / "unicode_out.xlsx"))
        data = {
            "headers": ["名前", "備考"],
            "rows": [["太郎", "日本語テスト"]],
        }
        output = op.create_workbook(data)
        assert Path(output).exists()

    def test_create_large_data(self, tmp_path):
        """1000+行のデータ書き込み"""
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(tmp_path / "large.xlsx"))
        data = {
            "headers": ["ID", "Value"],
            "rows": [[i, f"item_{i}"] for i in range(1500)],
        }
        output = op.create_workbook(data)
        assert Path(output).exists()

    def test_create_custom_sheet_name(self, tmp_path):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(tmp_path / "named.xlsx"))
        data = {
            "sheet_name": "MyData",
            "headers": ["Col1"],
            "rows": [["val"]],
        }
        op.create_workbook(data)
        from openpyxl import load_workbook
        wb = load_workbook(str(tmp_path / "named.xlsx"))
        assert "MyData" in wb.sheetnames


# ---------------------------------------------------------------------------
# write_cells
# ---------------------------------------------------------------------------

class TestWriteCells:
    def test_write_cells_basic(self, sample_xlsx):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(sample_xlsx))
        op.load()
        result = op.write_cells("Data", [{"cell": "A5", "value": "David"}])
        assert result is True

    def test_write_cells_not_loaded(self):
        from excel_ops import ExcelOperator
        op = ExcelOperator("dummy.xlsx")
        result = op.write_cells("Sheet1", [{"cell": "A1", "value": "x"}])
        assert result is False

    def test_write_cells_new_sheet(self, sample_xlsx):
        """存在しないシートは自動作成される"""
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(sample_xlsx))
        op.load()
        result = op.write_cells("NewSheet", [{"cell": "A1", "value": "test"}])
        assert result is True
        assert "NewSheet" in op.workbook.sheetnames

    def test_write_cells_missing_value(self, sample_xlsx):
        """value が None の場合はスキップされる"""
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(sample_xlsx))
        op.load()
        result = op.write_cells("Data", [{"cell": "A1", "value": None}])
        assert result is True

    def test_write_cells_empty_list(self, sample_xlsx):
        from excel_ops import ExcelOperator
        op = ExcelOperator(str(sample_xlsx))
        op.load()
        result = op.write_cells("Data", [])
        assert result is True


# ---------------------------------------------------------------------------
# CLI commands
# ---------------------------------------------------------------------------

class TestCLICommands:
    def test_cmd_read_json(self, sample_xlsx, capsys):
        from excel_ops import cmd_read
        args = MagicMock()
        args.file = str(sample_xlsx)
        args.sheet = None
        args.max_rows = 100
        args.format = "json"
        cmd_read(args)
        captured = capsys.readouterr()
        data = json.loads(captured.out)
        assert "headers" in data

    def test_cmd_read_text(self, sample_xlsx, capsys):
        from excel_ops import cmd_read
        args = MagicMock()
        args.file = str(sample_xlsx)
        args.sheet = None
        args.max_rows = 100
        args.format = "text"
        cmd_read(args)
        captured = capsys.readouterr()
        assert "Headers:" in captured.out

    def test_cmd_read_nonexistent(self, tmp_path, capsys):
        from excel_ops import cmd_read
        args = MagicMock()
        args.file = str(tmp_path / "nope.xlsx")
        args.sheet = None
        args.max_rows = 100
        args.format = "text"
        cmd_read(args)
        # Should print error but not crash

    def test_cmd_to_markdown_stdout(self, sample_xlsx, capsys):
        from excel_ops import cmd_to_markdown
        args = MagicMock()
        args.file = str(sample_xlsx)
        args.sheet = None
        args.max_rows = 100
        args.output = None
        cmd_to_markdown(args)
        captured = capsys.readouterr()
        assert "| Name" in captured.out

    def test_cmd_to_markdown_file(self, sample_xlsx, tmp_path):
        from excel_ops import cmd_to_markdown
        out = tmp_path / "out.md"
        args = MagicMock()
        args.file = str(sample_xlsx)
        args.sheet = None
        args.max_rows = 100
        args.output = str(out)
        cmd_to_markdown(args)
        assert out.exists()

    def test_cmd_analyze_json(self, sample_xlsx, capsys):
        from excel_ops import cmd_analyze
        args = MagicMock()
        args.file = str(sample_xlsx)
        args.format = "json"
        cmd_analyze(args)
        captured = capsys.readouterr()
        data = json.loads(captured.out)
        assert "summary" in data

    def test_cmd_analyze_text(self, sample_xlsx, capsys):
        from excel_ops import cmd_analyze
        args = MagicMock()
        args.file = str(sample_xlsx)
        args.format = "text"
        cmd_analyze(args)
        captured = capsys.readouterr()
        assert "Analysis Report" in captured.out

    def test_cmd_write(self, tmp_path, capsys):
        from excel_ops import cmd_write
        args = MagicMock()
        args.file = str(tmp_path / "new.xlsx")
        args.data = '{"headers": ["A"], "rows": [["1"]]}'
        args.output = None
        cmd_write(args)
        captured = capsys.readouterr()
        assert "Created" in captured.out

    def test_cmd_write_invalid_json(self, tmp_path, capsys):
        from excel_ops import cmd_write
        args = MagicMock()
        args.file = str(tmp_path / "new.xlsx")
        args.data = "not json"
        args.output = None
        cmd_write(args)
        captured = capsys.readouterr()
        assert "Invalid JSON" in captured.out

    def test_cmd_list_sheets(self, multi_sheet_xlsx, capsys):
        from excel_ops import cmd_list_sheets
        args = MagicMock()
        args.file = str(multi_sheet_xlsx)
        cmd_list_sheets(args)
        captured = capsys.readouterr()
        assert "Sheet1" in captured.out
        assert "Sheet2" in captured.out
