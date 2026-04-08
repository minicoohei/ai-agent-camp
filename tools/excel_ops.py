#!/usr/bin/env python3
"""
Excel Operations - Excel ファイル操作ツール

openpyxl を使用して Excel ファイルの読み取り・書き込み・分析を行います。

使用方法:
    uv run python tools/excel_ops.py read <file.xlsx>                  # 読み取り
    uv run python tools/excel_ops.py read <file.xlsx> --sheet "Sheet1" # 特定シート
    uv run python tools/excel_ops.py to-markdown <file.xlsx>           # Markdown変換
    uv run python tools/excel_ops.py analyze <file.xlsx>               # 分析レポート
    uv run python tools/excel_ops.py write <file.xlsx> --data '{...}'  # 書き込み
"""

import argparse
import json
import os
import sys
from pathlib import Path
from typing import Optional, Dict, List, Any
from datetime import datetime

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl library not found.")
    print("Please install with: pip install openpyxl")
    sys.exit(1)


class ExcelOperator:
    """Excel ファイル操作クラス"""
    
    def __init__(self, filepath: str):
        self.filepath = Path(filepath)
        self.workbook = None
    
    def load(self) -> bool:
        """ワークブックを読み込み"""
        if not self.filepath.exists():
            print(f"❌ File not found: {self.filepath}")
            return False
        
        try:
            self.workbook = load_workbook(str(self.filepath), data_only=True)
            return True
        except Exception as e:
            print(f"❌ Failed to load workbook: {e}")
            return False
    
    def get_sheet_names(self) -> List[str]:
        """シート名一覧を取得"""
        if not self.workbook:
            return []
        return self.workbook.sheetnames
    
    def read_sheet(self, sheet_name: str = None, max_rows: int = 1000) -> Dict[str, Any]:
        """シートのデータを読み取り"""
        if not self.workbook:
            return {"error": "Workbook not loaded"}
        
        if sheet_name:
            if sheet_name not in self.workbook.sheetnames:
                return {"error": f"Sheet '{sheet_name}' not found"}
            sheet = self.workbook[sheet_name]
        else:
            sheet = self.workbook.active
        
        data = {
            "sheet_name": sheet.title,
            "dimensions": sheet.dimensions,
            "max_row": min(sheet.max_row, max_rows),
            "max_col": sheet.max_column,
            "headers": [],
            "rows": []
        }
        
        # ヘッダー行を取得（1行目）
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            data["headers"].append(str(cell.value) if cell.value else f"Col{col}")
        
        # データ行を取得
        for row_num in range(2, min(sheet.max_row + 1, max_rows + 1)):
            row_data = []
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_num, column=col)
                value = cell.value
                if value is None:
                    value = ""
                elif isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    value = str(value)
                row_data.append(value)
            
            # 空行はスキップ
            if any(v for v in row_data):
                data["rows"].append(row_data)
        
        return data
    
    def to_markdown(self, sheet_name: str = None, max_rows: int = 100) -> str:
        """シートをMarkdownテーブルに変換"""
        data = self.read_sheet(sheet_name, max_rows)
        
        if "error" in data:
            return f"Error: {data['error']}"
        
        md_lines = []
        md_lines.append(f"# {self.filepath.name}")
        md_lines.append(f"\n**Sheet**: {data['sheet_name']}")
        md_lines.append(f"**Dimensions**: {data['dimensions']}")
        md_lines.append(f"**Rows**: {len(data['rows'])} (max {max_rows})")
        md_lines.append("")
        
        headers = data["headers"]
        if headers:
            md_lines.append("| " + " | ".join(headers) + " |")
            md_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
            
            for row in data["rows"]:
                # パイプ文字をエスケープ
                escaped_row = [str(v).replace("|", "\\|") for v in row]
                md_lines.append("| " + " | ".join(escaped_row) + " |")
        
        return "\n".join(md_lines)
    
    def analyze(self) -> Dict[str, Any]:
        """ワークブックの分析レポートを生成"""
        if not self.workbook:
            return {"error": "Workbook not loaded"}
        
        report = {
            "filename": self.filepath.name,
            "sheets": [],
            "total_cells": 0,
            "total_rows": 0,
            "summary": {}
        }
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            sheet_info = {
                "name": sheet_name,
                "dimensions": sheet.dimensions,
                "max_row": sheet.max_row,
                "max_col": sheet.max_column,
                "headers": [],
                "column_types": {}
            }
            
            # ヘッダー取得
            for col in range(1, min(sheet.max_column + 1, 50)):
                cell = sheet.cell(row=1, column=col)
                header = str(cell.value) if cell.value else f"Col{col}"
                sheet_info["headers"].append(header)
                
                # 列のデータ型を推測
                col_values = []
                for row in range(2, min(sheet.max_row + 1, 100)):
                    val = sheet.cell(row=row, column=col).value
                    if val is not None:
                        col_values.append(type(val).__name__)
                
                if col_values:
                    # 最も多いデータ型を取得
                    from collections import Counter
                    type_counts = Counter(col_values)
                    most_common = type_counts.most_common(1)[0][0]
                    sheet_info["column_types"][header] = most_common
            
            report["sheets"].append(sheet_info)
            report["total_cells"] += sheet.max_row * sheet.max_column
            report["total_rows"] += sheet.max_row
        
        report["summary"] = {
            "sheet_count": len(report["sheets"]),
            "estimated_cells": report["total_cells"],
            "total_rows": report["total_rows"]
        }
        
        return report
    
    def create_workbook(self, data: Dict[str, Any], output_path: str = None) -> str:
        """データからExcelファイルを作成"""
        wb = Workbook()
        ws = wb.active
        
        sheet_name = data.get("sheet_name", "Sheet1")
        ws.title = sheet_name
        
        headers = data.get("headers", [])
        rows = data.get("rows", [])
        
        # ヘッダー行を書き込み
        if headers:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
        
        # データ行を書き込み
        for row_num, row_data in enumerate(rows, 2):
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)
        
        # 列幅を自動調整
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            
            for cell in ws[column]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # 保存
        output = output_path or str(self.filepath)
        wb.save(output)
        
        return output
    
    def write_cells(self, sheet_name: str, updates: List[Dict]) -> bool:
        """
        セルを更新
        
        Args:
            sheet_name: シート名
            updates: 更新データのリスト [{"cell": "A1", "value": "xxx"}, ...]
        """
        if not self.workbook:
            return False
        
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)
        
        sheet = self.workbook[sheet_name]
        
        for update in updates:
            cell_ref = update.get("cell")
            value = update.get("value")
            
            if cell_ref and value is not None:
                sheet[cell_ref] = value
        
        self.workbook.save(str(self.filepath))
        return True


def cmd_read(args):
    """読み取りコマンド"""
    op = ExcelOperator(args.file)
    if not op.load():
        return
    
    data = op.read_sheet(sheet_name=args.sheet, max_rows=args.max_rows)
    
    if args.format == "json":
        print(json.dumps(data, ensure_ascii=False, indent=2))
    else:
        # 簡易表示
        print(f"Sheet: {data.get('sheet_name', 'N/A')}")
        print(f"Dimensions: {data.get('dimensions', 'N/A')}")
        print(f"Rows: {len(data.get('rows', []))}")
        print()
        print("Headers:", data.get("headers", []))
        print()
        print("Sample rows (first 5):")
        for i, row in enumerate(data.get("rows", [])[:5], 1):
            print(f"  {i}: {row}")


def cmd_to_markdown(args):
    """Markdown変換コマンド"""
    op = ExcelOperator(args.file)
    if not op.load():
        return
    
    md = op.to_markdown(sheet_name=args.sheet, max_rows=args.max_rows)
    
    if args.output:
        Path(args.output).write_text(md, encoding="utf-8")
        print(f"✅ Saved to: {args.output}")
    else:
        print(md)


def cmd_analyze(args):
    """分析コマンド"""
    op = ExcelOperator(args.file)
    if not op.load():
        return
    
    report = op.analyze()
    
    if args.format == "json":
        print(json.dumps(report, ensure_ascii=False, indent=2))
    else:
        print(f"📊 Analysis Report: {report['filename']}")
        print("=" * 50)
        print(f"Sheets: {report['summary']['sheet_count']}")
        print(f"Total rows: {report['summary']['total_rows']}")
        print(f"Estimated cells: {report['summary']['estimated_cells']}")
        print()
        
        for sheet in report["sheets"]:
            print(f"\n📋 {sheet['name']}")
            print(f"   Dimensions: {sheet['dimensions']}")
            print(f"   Rows: {sheet['max_row']}, Columns: {sheet['max_col']}")
            print(f"   Headers: {sheet['headers'][:10]}{'...' if len(sheet['headers']) > 10 else ''}")


def cmd_write(args):
    """書き込みコマンド"""
    try:
        data = json.loads(args.data)
    except json.JSONDecodeError as e:
        print(f"❌ Invalid JSON data: {e}")
        return
    
    op = ExcelOperator(args.file)
    output = op.create_workbook(data, output_path=args.output)
    print(f"✅ Created: {output}")


def cmd_list_sheets(args):
    """シート一覧コマンド"""
    op = ExcelOperator(args.file)
    if not op.load():
        return
    
    sheets = op.get_sheet_names()
    print(f"📋 Sheets in {args.file}:")
    for i, name in enumerate(sheets, 1):
        print(f"  {i}. {name}")


def main():
    parser = argparse.ArgumentParser(description="Excel Operations Tool")
    subparsers = parser.add_subparsers(dest="command", help="Commands")
    
    # read コマンド
    read_parser = subparsers.add_parser("read", help="Excelファイルを読み取り")
    read_parser.add_argument("file", help="Excelファイルパス")
    read_parser.add_argument("--sheet", "-s", help="シート名")
    read_parser.add_argument("--max-rows", "-n", type=int, default=100, help="最大行数")
    read_parser.add_argument("--format", "-f", choices=["text", "json"], default="text")
    
    # to-markdown コマンド
    md_parser = subparsers.add_parser("to-markdown", help="Markdownテーブルに変換")
    md_parser.add_argument("file", help="Excelファイルパス")
    md_parser.add_argument("--sheet", "-s", help="シート名")
    md_parser.add_argument("--max-rows", "-n", type=int, default=100)
    md_parser.add_argument("--output", "-o", help="出力ファイルパス")
    
    # analyze コマンド
    analyze_parser = subparsers.add_parser("analyze", help="ワークブックを分析")
    analyze_parser.add_argument("file", help="Excelファイルパス")
    analyze_parser.add_argument("--format", "-f", choices=["text", "json"], default="text")
    
    # write コマンド
    write_parser = subparsers.add_parser("write", help="Excelファイルを作成")
    write_parser.add_argument("file", help="出力ファイルパス")
    write_parser.add_argument("--data", "-d", required=True, help="JSON形式のデータ")
    write_parser.add_argument("--output", "-o", help="別ファイルに出力")
    
    # list-sheets コマンド
    list_parser = subparsers.add_parser("list-sheets", help="シート一覧を表示")
    list_parser.add_argument("file", help="Excelファイルパス")
    
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        sys.exit(1)
    
    if args.command == "read":
        cmd_read(args)
    elif args.command == "to-markdown":
        cmd_to_markdown(args)
    elif args.command == "analyze":
        cmd_analyze(args)
    elif args.command == "write":
        cmd_write(args)
    elif args.command == "list-sheets":
        cmd_list_sheets(args)


if __name__ == "__main__":
    main()
